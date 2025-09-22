import datetime
import calendar
import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

EXCEL_FILE = f"hodim_hisobot_{datetime.datetime.now().strftime('%Y-%m')}.xlsx"
ADMIN_ID = 5318613615

# Ish vaqti qoidalari
WORK_START = "08:10"
WORK_END = "20:00"
OVERTIME_GREEN_END = "21:00"
OVERTIME_YELLOW_START = "21:10"
VERY_LATE_LEAVE = "23:30"
EXTENDED_ARRIVAL_1 = "09:20"   # kecha 21:10+ bo'lsa ertalab shu vaqtdan kech hisob
EXTENDED_ARRIVAL_2 = "11:30"   # kecha 23:30+ bo'lsa ertalab shu vaqtdan kech hisob

# Jarimalar
FINE_PER_DAY = 20000            # sababsiz kech kelish
ABSENCE_FINE_PER_DAY = 200000   # kelmagan kun

user_status = {}

USER_COLORS = [
	"4F81BD", "9BBB59", "C0504D", "8064A2", "4BACC6", "F79646",
	"8FAADC", "A9D08E", "E26B5B", "B1A0C7", "76D6E2", "FABF8F",
	"BDD7EE", "C6E0B4", "F4B084", "ED7D31", "70AD47", "5B9BD5"
]

THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
					 top=Side(style="thin"), bottom=Side(style="thin"))

def _normalize_cell_date(cell_value) -> str:
	if isinstance(cell_value, datetime.datetime):
		return cell_value.strftime("%Y-%m-%d")
	if isinstance(cell_value, datetime.date):
		return cell_value.strftime("%Y-%m-%d")
	if cell_value is None:
		return ""
	return str(cell_value).split(" ")[0]

def _get_user_color(username: str) -> str:
	idx = abs(hash(username)) % len(USER_COLORS)
	return USER_COLORS[idx]

def _time_to_minutes(t: str) -> int:
	parts = t.strip().split(":")
	h = int(parts[0])
	m = int(parts[1])
	return h * 60 + m

def get_now_time_str():
	now = datetime.datetime.now()
	hour = str(int(now.strftime("%H")))
	minute = now.strftime("%M")
	return f"{hour}:{minute}"

def is_sunday(d: datetime.date) -> bool:
	return d.weekday() == 6  # Monday=0 ... Sunday=6

def style_date_headers(ws: Worksheet):
	a1 = ws.cell(row=1, column=1, value="Sana")
	a1.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
	a1.font = Font(color="FFFFFF", bold=True)
	a1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	a1.border = THIN_BORDER

	a2 = ws.cell(row=2, column=1, value="")
	a2.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
	a2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	a2.border = THIN_BORDER

	ws.freeze_panes = "B3"

def style_date_column(ws: Worksheet):
	weekday_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
	sunday_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
	date_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for row_idx in range(3, ws.max_row + 1):
		cell = ws.cell(row=row_idx, column=1)
		d = None
		raw = cell.value
		try:
			if isinstance(raw, datetime.datetime):
				d = raw.date()
			elif isinstance(raw, datetime.date):
				d = raw
			elif raw:
				d = datetime.datetime.strptime(str(raw).split(" ")[0], "%Y-%m-%d").date()
		except Exception:
			d = None
		if d and is_sunday(d):
			cell.fill = sunday_fill
			cell.font = Font(bold=True)
		else:
			cell.fill = weekday_fill
		cell.alignment = date_align
		cell.border = THIN_BORDER

def apply_user_header_style(ws: Worksheet, start_col: int, user_name: str):
	end_col = start_col + 1
	ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
	header_cell = ws.cell(row=1, column=start_col)
	user_color = _get_user_color(user_name)
	header_cell.value = user_name
	header_cell.fill = PatternFill(start_color=user_color, end_color=user_color, fill_type="solid")
	header_cell.font = Font(color="FFFFFF", bold=True)
	header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	header_cell.border = THIN_BORDER

	in_cell = ws.cell(row=2, column=start_col, value="Keldim")
	out_cell = ws.cell(row=2, column=end_col, value="Ketyapman")
	for cell in (in_cell, out_cell):
		cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = THIN_BORDER

def tint_user_column(ws: Worksheet, start_col: int, user_name: str):
	base = _get_user_color(user_name)
	light_map = {
		"4F81BD":"D9E3F2","9BBB59":"E6F1DA","C0504D":"F4DCDB","8064A2":"E8E2EF","4BACC6":"D8EEF4","F79646":"FCE8D9",
		"8FAADC":"E9F0FB","A9D08E":"EDF5E6","E26B5B":"F8E1DE","B1A0C7":"EEEAF4","76D6E2":"E7F8FA","FABF8F":"FEF0E3",
		"BDD7EE":"EFF6FD","C6E0B4":"F0F7EA","F4B084":"FDEFE5","ED7D31":"FBE4D6","70AD47":"EAF4E3","5B9BD5":"E7F0FA"
	}
	light = light_map.get(base, "F5F5F5")
	fill = PatternFill(start_color=light, end_color=light, fill_type="solid")
	align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for row_idx in range(3, ws.max_row + 1):
		for col in (start_col, start_col + 1):
			cell = ws.cell(row=row_idx, column=col)
			if cell.value in (None, ""):
				cell.fill = fill
			if not cell.alignment or cell.alignment.horizontal != "center":
				cell.alignment = align
			cell.border = THIN_BORDER

def autosize_columns(ws: Worksheet, min_width=10, max_width=22):
	ws.column_dimensions[get_column_letter(1)].width = 12
	for col_idx in range(2, ws.max_column + 1):
		column_letter = get_column_letter(col_idx)
		max_length = 0
		for row_idx in (1, 2):
			val = ws.cell(row=row_idx, column=col_idx).value
			if val:
				max_length = max(max_length, len(str(val)))
		for row_idx in range(3, ws.max_row + 1):
			val = ws.cell(row=row_idx, column=col_idx).value
			if val:
				max_length = max(max_length, len(str(val)))
		width = max(min_width, min(max_width, int(max_length * 1.2) + 2))
		ws.column_dimensions[column_letter].width = width

def create_excel():
	if not os.path.exists(EXCEL_FILE):
		wb = openpyxl.Workbook()
		ws = wb.active
		style_date_headers(ws)

		year = datetime.datetime.now().year
		month = datetime.datetime.now().month
		days_in_month = calendar.monthrange(year, month)[1]
		for i in range(1, days_in_month + 1):
			sana = datetime.date(year, month, i).strftime("%Y-%m-%d")
			ws.cell(row=i + 2, column=1, value=sana).border = THIN_BORDER

		style_date_column(ws)
		autosize_columns(ws)
		wb.save(EXCEL_FILE)

def _find_user_start_col(ws: Worksheet, user_name: str):
	for col in range(2, ws.max_column + 1):
		val = ws.cell(row=1, column=col).value
		if val == user_name:
			return col
	return None

def add_hodim_if_not_exists(user_name):
	wb = openpyxl.load_workbook(EXCEL_FILE)
	ws = wb.active

	start_col = _find_user_start_col(ws, user_name)
	if start_col is None:
		new_start_col = ws.max_column + 1
		apply_user_header_style(ws, new_start_col, user_name)
		for row in range(3, ws.max_row + 1):
			ws.cell(row=row, column=new_start_col, value="").border = THIN_BORDER
			ws.cell(row=row, column=new_start_col + 1, value="").border = THIN_BORDER
		tint_user_column(ws, new_start_col, user_name)
		autosize_columns(ws)
		wb.save(EXCEL_FILE)
	else:
		apply_user_header_style(ws, start_col, user_name)
		tint_user_column(ws, start_col, user_name)
		autosize_columns(ws)
		wb.save(EXCEL_FILE)

	wb.close()

def _ensure_today_row(ws: Worksheet, today_sana: str) -> int:
	for row in range(3, ws.max_row + 1):
		if _normalize_cell_date(ws.cell(row=row, column=1).value) == today_sana:
			return row
	ws.append([today_sana] + [""] * (ws.max_column - 1))
	new_row = ws.max_row
	date_cell = ws.cell(row=new_row, column=1)
	try:
		d = datetime.datetime.strptime(today_sana, "%Y-%m-%d").date()
	except Exception:
		d = None
	if d and is_sunday(d):
		date_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
		date_cell.font = Font(bold=True)
	else:
		date_cell.fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
	date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	date_cell.border = THIN_BORDER
	for col in range(2, ws.max_column + 1):
		ws.cell(row=new_row, column=col).border = THIN_BORDER
	return new_row

def _get_prev_day_leave(ws: Worksheet, start_col: int, today_sana: str):
	try:
		today = datetime.datetime.strptime(today_sana, "%Y-%m-%d").date()
	except Exception:
		return None
	prev = today - datetime.timedelta(days=1)
	prev_sana = prev.strftime("%Y-%m-%d")
	for row in range(3, ws.max_row + 1):
		if _normalize_cell_date(ws.cell(row=row, column=1).value) == prev_sana:
			val = ws.cell(row=row, column=start_col + 1).value
			return None if val in (None, "") else str(val)
	return None

def _dynamic_arrival_limit(prev_leave_time: str | None) -> int:
	limit_minutes = _time_to_minutes(WORK_START)
	if prev_leave_time:
		prev_m = _time_to_minutes(prev_leave_time)
		if prev_m >= _time_to_minutes(VERY_LATE_LEAVE):
			limit_minutes = _time_to_minutes(EXTENDED_ARRIVAL_2)  # 11:30
		elif prev_m >= _time_to_minutes(OVERTIME_YELLOW_START):
			limit_minutes = _time_to_minutes(EXTENDED_ARRIVAL_1)  # 09:20
	return limit_minutes

def _color_arrival_cell(cell, arrival_time: str, prev_leave_time: str | None):
	if not arrival_time:
		return
	limit_minutes = _dynamic_arrival_limit(prev_leave_time)
	arr_m = _time_to_minutes(arrival_time)
	if arr_m <= limit_minutes:
		cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # green
	else:
		cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # red-ish

def _color_leave_cell(cell, leave_time: str):
	if not leave_time:
		return
	t_out = _time_to_minutes(leave_time)
	if t_out < _time_to_minutes(WORK_END):
		cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # red-ish
	elif _time_to_minutes(WORK_END) <= t_out <= _time_to_minutes(OVERTIME_GREEN_END):
		cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # green
	elif t_out >= _time_to_minutes(OVERTIME_YELLOW_START):
		cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # yellow

def write_user_time(user_name: str, status: str, time: str):
	wb = openpyxl.load_workbook(EXCEL_FILE)
	ws = wb.active
	today_sana = datetime.datetime.now().strftime("%Y-%m-%d")

	start_col = _find_user_start_col(ws, user_name)
	if start_col is None:
		add_hodim_if_not_exists(user_name)
		wb.close()
		wb = openpyxl.load_workbook(EXCEL_FILE)
		ws = wb.active
		start_col = _find_user_start_col(ws, user_name)
		if start_col is None:
			wb.close()
			return

	row_idx = _ensure_today_row(ws, today_sana)

	arrival_cell = ws.cell(row=row_idx, column=start_col)
	leave_cell = ws.cell(row=row_idx, column=start_col + 1)
	existing_in = "" if arrival_cell.value in (None, "") else str(arrival_cell.value).strip()
	existing_out = "" if leave_cell.value in (None, "") else str(leave_cell.value).strip()

	if status == "keldi" and not existing_in:
		arrival_cell.value = time
	if status == "ketdi" and not existing_out:
		leave_cell.value = time

	align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for cell in (arrival_cell, leave_cell):
		cell.alignment = align
		cell.border = THIN_BORDER

	prev_leave = _get_prev_day_leave(ws, start_col, today_sana)
	_color_arrival_cell(arrival_cell, arrival_cell.value if arrival_cell.value else "", prev_leave)
	_color_leave_cell(leave_cell, leave_cell.value if leave_cell.value else "")

	apply_user_header_style(ws, start_col, user_name)
	tint_user_column(ws, start_col, user_name)
	autosize_columns(ws)

	wb.save(EXCEL_FILE)
	wb.close()

def build_keyboard(user_id: int):
	if user_id == ADMIN_ID:
		reply_keyboard = [
			[KeyboardButton("Keldim"), KeyboardButton("Ketyapman")],
			[KeyboardButton("Oylik hisobot")]
		]
	else:
		reply_keyboard = [
			[KeyboardButton("Keldim"), KeyboardButton("Ketyapman")]
		]
	return ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True, one_time_keyboard=True)

def get_cycle_bounds_15_to_15(today: datetime.date | None = None) -> tuple[datetime.date, datetime.date]:
	if today is None:
		today = datetime.datetime.now().date()
	y, m, d = today.year, today.month, today.day
	if d >= 15:
		start = datetime.date(y, m, 15)
		ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
		end = datetime.date(ny, nm, 14)
	else:
		py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
		start = datetime.date(py, pm, 15)
		end = datetime.date(y, m, 14)
	return start, end

def _get_date_from_row(ws, row_idx: int):
	raw = ws.cell(row=row_idx, column=1).value
	if not raw:
		return None
	try:
		if isinstance(raw, datetime.datetime):
			return raw.date()
		if isinstance(raw, datetime.date):
			return raw
		parts = str(raw).split(" ")[0]
		return datetime.datetime.strptime(parts, "%Y-%m-%d").date()
	except Exception:
		return None

def _parse_hhmm(value):
	if not value:
		return ""
	return str(value).strip()

def _is_unexcused_late(arrival_hhmm: str, prev_leave_hhmm: str | None) -> bool:
	if not arrival_hhmm:
		return False
	limit_minutes = _dynamic_arrival_limit(prev_leave_hhmm)
	return _time_to_minutes(arrival_hhmm) > limit_minutes

def mark_absence(ws, row_idx: int, start_col: int):
	black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	white_bold = Font(color="FFFFFF", bold=True)
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for c in (start_col, start_col + 1):
		cell = ws.cell(row=row_idx, column=c)
		cell.value = "200 000"
		cell.fill = black
		cell.font = white_bold
		cell.alignment = center
		cell.border = THIN_BORDER

def compute_and_write_fines_summary(excel_path: str):
	wb = openpyxl.load_workbook(excel_path)
	ws = wb.active

	start_date, end_date = get_cycle_bounds_15_to_15()

	# Foydalanuvchilar (har biri 2 ustun)
	users = []
	col = 2
	while col <= ws.max_column:
		name = ws.cell(row=1, column=col).value
		if name:
			users.append((name, col))
			col += 2
		else:
			col += 1

	# Sana satrlari (yakshanba SKIP)
	date_rows = []
	for r in range(3, ws.max_row + 1):
		d = _get_date_from_row(ws, r)
		if d and (start_date <= d <= end_date) and not is_sunday(d):
			date_rows.append((r, d))

	user_to_late_days = {u: 0 for u, _ in users}
	user_to_absent_days = {u: 0 for u, _ in users}

	for user_name, start_col in users:
		for r, d in date_rows:
			arrival = _parse_hhmm(ws.cell(row=r, column=start_col).value)
			leave = _parse_hhmm(ws.cell(row=r, column=start_col + 1).value)

			# Kelmagan — ikkisi ham bo'sh
			if not arrival and not leave:
				user_to_absent_days[user_name] += 1
				mark_absence(ws, r, start_col)
				continue

			# Sababsiz kechikish (dinamik limit)
			prev_row = None
			prev_date = d - datetime.timedelta(days=1)
			for rr in range(3, ws.max_row + 1):
				dd = _get_date_from_row(ws, rr)
				if dd == prev_date:
					prev_row = rr
					break
			prev_leave = None
			if prev_row:
				prev_leave = _parse_hhmm(ws.cell(row=prev_row, column=start_col + 1).value)

			if _is_unexcused_late(arrival, prev_leave):
				user_to_late_days[user_name] += 1

	# Hisobot varaq (Excel)
	sheet_name = "Hisobot"
	if sheet_name in wb.sheetnames:
		del wb[sheet_name]
	rep = wb.create_sheet(title=sheet_name)

	rep.cell(row=1, column=1, value=f"Hisob-kitob davri: {start_date.strftime('%Y-%m-%d')} → {end_date.strftime('%Y-%m-%d')} (15→15), yakshanba — dam (jarimasiz)")
	rep.cell(row=3, column=1, value="Hodim")
	rep.cell(row=3, column=2, value="Sababsiz kech kelgan (kun)")
	rep.cell(row=3, column=3, value=f"Jarima (kech) [x {FINE_PER_DAY}]")
	rep.cell(row=3, column=4, value="Kelmagan (kun)")
	rep.cell(row=3, column=5, value=f"Jarima (kelmagan) [x {ABSENCE_FINE_PER_DAY}]")
	rep.cell(row=3, column=6, value="Jami jarima")

	header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	for c in range(1, 7):
		h = rep.cell(row=3, column=c)
		h.fill = header_fill
		h.font = header_font
		h.alignment = Alignment(horizontal="center", vertical="center")

	row = 4
	for user_name, _ in users:
		late_days = user_to_late_days[user_name]
		absent_days = user_to_absent_days[user_name]
		fine_late = late_days * FINE_PER_DAY
		fine_abs = absent_days * ABSENCE_FINE_PER_DAY
		total = fine_late + fine_abs

		rep.cell(row=row, column=1, value=user_name)
		rep.cell(row=row, column=2, value=late_days)
		rep.cell(row=row, column=3, value=fine_late)
		rep.cell(row=row, column=4, value=absent_days)
		rep.cell(row=row, column=5, value=fine_abs)
		rep.cell(row=row, column=6, value=total)
		for c in range(1, 7):
			rep.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
		row += 1

	rep.column_dimensions[get_column_letter(1)].width = 28
	for c in range(2, 7):
		rep.column_dimensions[get_column_letter(c)].width = 22

	wb.save(excel_path)
	wb.close()
	return start_date, end_date

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
	user_id = update.message.from_user.id
	user_name = update.message.from_user.full_name
	create_excel()
	add_hodim_if_not_exists(user_name)
	markup = build_keyboard(user_id)
	await update.message.reply_text(
		f"Assalomu alaykum {user_name}!\nQuyidagi tugmalardan birini tanlang:",
		reply_markup=markup
	)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
	user_id = update.message.from_user.id
	user_name = update.message.from_user.full_name
	text = (update.message.text or "").strip()
	markup = build_keyboard(user_id)
	create_excel()
	add_hodim_if_not_exists(user_name)

	if text == "Keldim":
		user_status[user_id] = "keldi"
		await update.message.reply_text("Video dalil yuboring:", reply_markup=markup)
		return

	if text == "Ketyapman":
		user_status[user_id] = "ketdi"
		await update.message.reply_text("Video dalil yuboring:", reply_markup=markup)
		return

	if text == "Oylik hisobot" and user_id == ADMIN_ID:
		if os.path.exists(EXCEL_FILE):
			# Excel hisobotni yangilaymiz (15→15, yakshanba dam)
			start_date, end_date = compute_and_write_fines_summary(EXCEL_FILE)
			with open(EXCEL_FILE, "rb") as fx:
				await update.message.reply_document(document=fx, caption=f"Hisobot (Excel) {start_date} → {end_date}")
		else:
			await update.message.reply_text("Hisobot fayli topilmadi.")
		return

	if text == "Oylik hisobot" and user_id != ADMIN_ID:
		await update.message.reply_text("Bu tugma faqat admin uchun!", reply_markup=markup)
		return

	# Video yoki video_note
	if update.message.video or update.message.video_note:
		status = user_status.get(user_id)
		now = get_now_time_str()
		if status in ["keldi", "ketdi"]:
			write_user_time(user_name, status, now)
			user_status[user_id] = None
		if update.message.video:
			video_file_id = update.message.video.file_id
			await context.bot.send_video(
				chat_id=ADMIN_ID,
				video=video_file_id,
				caption=f"{user_name} {status if status else ''} {now}"
			)
		elif update.message.video_note:
			video_file_id = update.message.video_note.file_id
			await context.bot.send_video_note(
				chat_id=ADMIN_ID,
				video_note=video_file_id
			)
		await update.message.reply_text(
			f"Video adminga yuborildi. Vaqt: {now}",
			reply_markup=markup
		)
		return

	await update.message.reply_text("Iltimos, tugmalardan birini tanlang yoki video yuboring.", reply_markup=markup)

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
	try:
		print("Exception in handler:", context.error)
	except Exception:
		pass

if __name__ == "__main__":
	TOKEN = "8334665305:AAFykh9AZ1d4wgmlze_b8kx5rk2XvgRRCCA"
	create_excel()
	app = ApplicationBuilder().token(TOKEN).build()
	app.add_handler(CommandHandler("start", start))
	app.add_handler(MessageHandler(filters.ALL, handle_message))
	app.add_error_handler(error_handler)
	print("Bot ishga tushdi ✅")
	app.run_polling()